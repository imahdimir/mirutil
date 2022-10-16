"""

    """

from pathlib import Path

import openpyxl as pyxl
import pandas as pd


def save_df_as_a_nice_xl(df ,
                         fpn ,
                         index: bool = False ,
                         header: bool = True ,
                         max_col_length: int = 40) -> None :
    df.to_excel(fpn , index = False)

    wb = pyxl.load_workbook(fpn)
    ws = wb.active

    panes = index * ws['A'] + header * ws[1]

    for cell in panes :
        cell.style = 'Pandas'

    for column in ws.columns :
        length = max(len(str(cell.value)) for cell in column) + 3
        length = length if length <= max_col_length else max_col_length
        ws.column_dimensions[column[0].column_letter].width = length

    ws.freeze_panes = 'A2'

    wb.save(fpn)
    wb.close()
    print(f"saved as {fpn}")


def save_as_prq_wo_index(df , fpn) -> None :
    df.to_parquet(fpn , index = False)
    print(f'dataframe saved as {fpn} without index')


def read_data_according_to_type(fpn) -> pd.DataFrame :
    suf = Path(fpn).suffix
    if suf == '.xlsx' :
        return pd.read_excel(fpn)
    elif suf == '.prq' :
        return pd.read_parquet(fpn)
    elif suf == '.csv' :
        return pd.read_csv(fpn)


def has_extra_data(df , df1) -> bool :
    df = df.convert_dtypes()
    df1 = df1.convert_dtypes()

    df = df.astype('string')
    df1 = df1.astype('string')

    _df = pd.concat([df , df1])

    _df = _df.drop_duplicates()

    return not df.equals(_df)
