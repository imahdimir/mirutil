##


def convert_digits_to_en(istr) :
  from persiantools import digits


  if not isinstance(istr , str) :
    return istr

  os = digits.ar_to_fa(istr)
  os = digits.fa_to_en(os)

  return os

def rm_odd_chars(istr) :
  import re


  if not isinstance(istr , str) :
    return istr

  repmap = {
      r"\u202b" : ' ' ,
      r'\u200c' : ' ' ,
      r'\u200d' : '' ,
      r'\u200f' : '' ,
      }

  os = istr
  for ptr , rep in repmap.items() :
    os = re.sub(ptr , rep , istr)

  return os

def strip_and_rm_successive_spaces_in_between(istr) :
  import re


  if not isinstance(istr , str) :
    return istr

  os = re.sub('\s+' , ' ' , istr)
  os = os.strip()

  return os

def normalize_fa_str(fa_str: str) -> str :
  """ Normalize Persian/Farsi strings to a much simpler form for unification purposes

  Usage::
  >>> from mirutil import normalize as norm
  >>> converted = norm("آگاه نیکو")

  :param fa_str: A string, will be simplified
  :rtype: str
  """
  from persiantools import characters


  if not isinstance(fa_str , str) :
    return fa_str

  os = convert_digits_to_en(fa_str)
  os = characters.ar_to_fa(os)
  os = rm_odd_chars(os)
  os = strip_and_rm_successive_spaces_in_between(os)

  return os

def save_df_as_a_nice_xl(df ,
                         fpn ,
                         index: bool = False ,
                         header: bool = True ,
                         max_col_length: int = 40
                         ) :
  import openpyxl as pyxl


  df.to_excel(fpn , index = False)

  wb = pyxl.load_workbook(fpn)
  ws = wb.active

  panes = index * ws['A'] + header * ws[1]

  for cell in panes :
    cell.style = 'Pandas'

  for column in ws.columns :
    length = max(len(str(cell.value)) for cell in column) + 3
    length = length if length <= max_col_length else max_col_length
    ws.column_dimensions[column[0].column_letter].width = length

  ws.freeze_panes = 'A2'

  wb.save(fpn)
  wb.close()

  print(f"saved as {fpn}")

def save_as_prq_wo_index(df , fpn) -> None :
  df.to_parquet(fpn , index = False)
  print(f'dataframe saved as {fpn} without index')

def read_data_according_to_type(fpn) :
  from pathlib import Path
  import pandas as pd


  suf = Path(fpn).suffix
  if suf == '.xlsx' :
    return pd.read_excel(fpn)
  elif suf == '.prq' :
    return pd.read_parquet(fpn)
  elif suf == '.csv' :
    return pd.read_csv(fpn)

def update_metadata_save_rand_sample(fp , save_rand_sample = True) -> None :
  """
  :param fp:
  :param save_rand_sample:
  :return:
  """

  import json
  from pathlib import Path

  from .namespaces import MetadataColumns


  cns = MetadataColumns()

  dirpth = Path(fp).parent
  metafp = dirpth / 'META.json'
  if not metafp.exists() :
    return None

  with open(metafp) as fi :
    meta = json.load(fi)

  df = read_data_according_to_type(fp)

  if cns.startendcol in meta.keys() :
    if meta[cns.startendcol] is not None :
      meta[cns.start] = df[meta[cns.startendcol]].min()
      meta[cns.end] = df[meta[cns.startendcol]].max()

  meta[cns.numrow] = len(df)
  meta[cns.numcol] = len(df.columns)
  meta[cns.colnames] = list(df.columns)

  with open(metafp , 'w' , encoding = 'utf-8') as fi :
    json.dump(meta , fi , ensure_ascii = False)
    print("Meta updated.")

  if save_rand_sample and len(df) > 1000 :
    _df = df.sample(n = 1000)
    _fp = Path(fp).with_stem('Sample').with_suffix('.xlsx')
    save_df_as_a_nice_xl(_df , _fp)
    print('random sample saved.')

def persian_tools_jdate_from_iso_format_jdate_str(jdate_str: str) :
  import re

  import pandas as pd
  from persiantools.jdatetime import JalaliDate


  iso_fmt_jd = r'1[34]\d\d-[0-2]\d-[0-3]\d'

  if pd.isna(jdate_str) :
    return None

  jd = str(jdate_str)

  cnd = re.fullmatch(iso_fmt_jd , jd)

  if cnd is not None :
    return JalaliDate(int(jd[:4]) , int(jd[5 :7]) , int(jd[8 :10]))
  elif cnd is None :
    raise ValueError

def persian_tools_jdate_from_int_format_jdate(jdate: {int , str}) :
  import re

  import pandas as pd
  from persiantools.jdatetime import JalaliDate


  int_fmt_jd = r'1[34]\d\d[0-2]\d[0-3]\d'

  if pd.isna(jdate) :
    return None

  jd = str(int(jdate))

  cnd = re.fullmatch(int_fmt_jd , jd)

  if cnd is not None :
    return JalaliDate(int(jd[:4]) , int(jd[4 :6]) , int(jd[6 :8]))
  elif cnd is None :
    raise ValueError

def print_df_columns_in_dict_type(df) :
  for cn in df.columns :
    print('"' + cn + '":None,')

def extract_market_from_tsetmc_title(title: str) :
  import re


  os = title.replace("',FaraDesc ='" , ' ')
  os = os.strip()

  ptr = r'.+-\s*([^-]*)'
  if re.fullmatch(ptr , os) :
    return re.sub(ptr , r'\1' , os).strip()
  else :
    ptr = r'.+-\s*-\s*'
    assert re.fullmatch(ptr , os)

def search_tsetmc(string) :
  import requests
  import pandas as pd


  url = f'http://www.tsetmc.com/tsev2/data/search.aspx?skey={string}'

  order_map = {
      'Ticker'   : 0 ,
      'Name'     : 1 ,
      'ID-1'     : 2 ,
      'ID-2'     : 3 ,
      'ID-3'     : 4 ,
      'ID-4'     : 5 ,
      'unk6'     : 6 ,
      'IsActive' : 7 ,
      'unk8'     : 8 ,
      'unk9'     : 9 ,
      'Market'   : 10 ,
      }

  headers = {
      'User-Agent' : 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'
      }

  rsp = requests.get(url , headers = headers)

  rows = rsp.text.split(';')
  rows = [x for x in rows if x != '']

  df = pd.DataFrame(columns = order_map.keys())

  for rw in rows :
    vals = rw.split(',')
    dfr = order_map.copy()

    for ky , vl in order_map.items() :
      try :
        dfr[ky] = vals[vl]
      except IndexError :
        dfr[ky] = None

    _df = pd.DataFrame(data = dfr , index = [0])
    df = pd.concat([df , _df] , ignore_index = True)

  return df

def return_clusters_indices(iterable_obj , cluster_size = 100) :
  intdiv = len(iterable_obj) // cluster_size

  cis = [x * cluster_size for x in range(0 , intdiv + 1)]

  if len(cis) > 1 :
    if cis[-1] != len(iterable_obj) :
      cis.append(cis[-1] + len(iterable_obj) % cluster_size)
  else :
    cis = [0 , len(iterable_obj)]
    if cis == [0 , 0] :
      cis = [0]

  cis[0] = cis[0]

  se_tuples = []
  for _i in range(len(cis) - 1) :
    si = cis[_i]
    ei = cis[_i + 1] - 1
    se = (si , ei)
    se_tuples.append(se)

  print(se_tuples)
  return se_tuples

def get_an_id_testmc_overview_page_resp(tsetmc_id):
  import requests


  hdrs = {
      'User-Agent' : 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'
      }

  url = f'http://www.tsetmc.com/Loader.aspx?ParTree=151311&i={tsetmc_id}'

  resp = requests.get(url , headers = hdrs)

  return resp
def get_title_stocks_from_overview_page_by_stock_id(tsetmc_id) :
  import re


  resp = get_an_id_testmc_overview_page_resp(tsetmc_id)

  title_list = re.findall(r"Title='(.+)',FaraDesc" , resp.text)
  print(title_list)

  return title_list

def get_groupname_from_overview_page_by_stock_id(tsetmc_id):
  import re


  resp = get_an_id_testmc_overview_page_resp(tsetmc_id)

  gpns = re.findall(r"LSecVal='(.+)',CgrValCot" , resp.text)
  print(gpns)

  return gpns

def make_zero_padded_jdate(ist , sep = '/') :
  spl = ist.split(sep)

  for _i in range(1 , 3) :
    if int(spl[_i]) < 10 :
      spl[_i] = '0' + spl[_i]

  ou = '-'.join(spl)

  return ou

##


##
##