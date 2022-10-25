"""

    """

from itertools import product
from pathlib import Path

import openpyxl as pyxl
import pandas as pd

from typing import Union


def save_df_as_a_nice_xl(df ,
                         fpn ,
                         index: bool = False ,
                         header: bool = True ,
                         max_col_length: int = 40) -> None :
    df.to_excel(fpn , index = False)

    wb = pyxl.load_workbook(fpn)
    ws = wb.active

    panes = index * ws['A'] + header * ws[1]

    for cell in panes :
        cell.style = 'Pandas'

    for column in ws.columns :
        length = max(len(str(cell.value)) for cell in column) + 3
        length = length if length <= max_col_length else max_col_length
        ws.column_dimensions[column[0].column_letter].width = length

    ws.freeze_panes = 'A2'

    wb.save(fpn)
    wb.close()
    print(f"saved as {fpn}")

def update_with_last_run_data(df , fp) :
    if Path(fp).exists() :
        lastdf = pd.read_parquet(fp)
        df.update(lastdf)
    return df

def save_as_prq_wo_index(df , fp) -> None :
    df.to_parquet(fp , index = False)
    print(f'dataframe saved as {fp} without index')

def read_data_according_to_type(fp) -> pd.DataFrame :
    suf = Path(fp).suffix
    if suf == '.xlsx' :
        return pd.read_excel(fp)
    elif suf == '.prq' :
        return pd.read_parquet(fp)
    elif suf == '.csv' :
        return pd.read_csv(fp)

def has_extra_data(df , df1) -> bool :
    df = df.convert_dtypes()
    df1 = df1.convert_dtypes()

    df = df.astype('string')
    df1 = df1.astype('string')

    _df = pd.concat([df , df1])

    _df = _df.drop_duplicates()

    return not df.equals(_df)

def drop_dup_and_sub_dfs(dfs_list: list) -> list :
    """

    assumption: len(dfs_list) >= 2
    """

    dfs1 = [(i , df) for i , df in enumerate(dfs_list)]

    pr = product(dfs_list , dfs1)
    pr = list(pr)

    df = pd.DataFrame(pr , columns = [0 , 1])

    df[2] = df[1].apply(lambda x : x[0])
    df[1] = df[1].apply(lambda x : x[1])

    ms = df.apply(lambda x : x[0].equals(x[1]) , axis = 1)
    df = df[~ ms]

    df[3] = df.apply(lambda x : has_extra_data(x[0] , x[1]) , axis = 1)

    df[4] = df.groupby(2)[3].transform('all')

    ms = df[4]
    df = df[ms]

    df = df.drop_duplicates(subset = 2)

    return df[1].to_list()

def find_all_df_locs_eq_val(df: pd.DataFrame , val) -> pd.MultiIndex :
    msk = df.eq(val)
    _df = df[msk]
    s = _df.stack()
    return s.index

def ret_north_west_of_indices(mi: pd.MultiIndex) :
    df = mi.to_frame()
    if df.empty :
        return
    df = df.sort_values(by = [0 , 1] , ascending = False)
    r = df.iloc[[0]]
    return r.index[0]

def ret_mask_indices(msk) :
    return msk[msk].index

def ret_indices(df , msk = None) :
    if msk is not None :
        return ret_mask_indices(msk)
    else :
        return df.index

def make_inputs_4_apply_parallel(df , inds , inp_cols) :
    if len(inp_cols) == 1 :
        return df.loc[inds , inp_cols[0]]
    else :
        df = df.loc[inds , inp_cols]
        return zip(*[df[x].to_list() for x in inp_cols])

def run_parallel(func , inp , n_jobs = 30 , console_run = True) :
    if console_run :
        from multiprocess.pool import Pool
    else :
        from multiprocessing import Pool

    pool = Pool(n_jobs)

    if isinstance(inp , zip) :
        fu = pool.starmap
    else :
        fu = pool.map

    try :
        return fu(func , inp)
    except KeyboardInterrupt :
        return

def handle_parallel_output(o , df , inds , out_map) :
    if isinstance(out_map , dict) :
        for k , v in out_map.items() :
            df.loc[inds , k] = [x.__getattribute__(v) for x in o]

    elif isinstance(out_map , list) :
        for i , col in enumerate(out_map) :
            df.loc[inds , col] = [x[i] for x in o]

    return df

def make_out_map_ready(out_cols_map) :
    if out_cols_map is None :
        return

    if isinstance(out_cols_map , dict) :

        if list(out_cols_map.values())[0] is None :
            return list(out_cols_map.keys())
        else :
            return out_cols_map

    elif isinstance(out_cols_map , list) :
        return out_cols_map

def df_apply_parallel(df ,
                      func ,
                      inp_cols: list ,
                      out_cols_map: Union[list , dict , None] = None ,
                      msk = None ,
                      test = False ,
                      n_jobs = 32 ,
                      console_run = True) :

    inds = ret_indices(df , msk)
    if test :
        inds = inds[: min(100 , len(inds))]

    inp = make_inputs_4_apply_parallel(df , inds , inp_cols)

    o = run_parallel(func , inp , n_jobs , console_run = console_run)

    ocm = make_out_map_ready(out_cols_map)
    df = handle_parallel_output(o , df , inds , ocm)

    return df
